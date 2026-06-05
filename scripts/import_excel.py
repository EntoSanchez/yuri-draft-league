#!/usr/bin/env python3
"""Import season archives from Excel files into league.db.

Usage:
    python scripts/import_excel.py [6] [7] [8]   # import specific seasons
    python scripts/import_excel.py                # import all seasons

Excel files are resolved in this order:
  1. PROJECT_ROOT/data/<filename>   ← upload files here on PythonAnywhere
  2. ~/Downloads/<filename>         ← Windows dev machine fallback

Each season is upserted by season_num — re-running is safe.
"""

import json
import os
import re
import sys
import sqlite3
from datetime import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not found — run: pip install openpyxl")

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.environ.get("LEAGUE_DB", os.path.join(PROJECT_ROOT, "league.db"))


def _resolve_file(filename):
    """Find the Excel file: project data/ dir first, then ~/Downloads."""
    candidates = [
        os.path.join(PROJECT_ROOT, "data", filename),
        os.path.join(os.path.expanduser("~"), "Downloads", filename),
        # Windows dev machine full path kept as last-resort
        os.path.join(r"C:\Users\zcs55\Downloads", filename),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return candidates[0]  # will fail with a clear error below


SEASONS = {
    8: {
        "name": "Yuri Cup Season 8",
        "file": _resolve_file("Yuri Cup Season 8 For Real (2).xlsx"),
        "sheet": "Data",
        "fmt": "data_sheet",
    },
    7: {
        "name": "Yuri Cup Season 7",
        "file": _resolve_file("Yuri Cup Season 7 (1).xlsx"),
        "sheet": "Data",
        "fmt": "data_sheet",
    },
    6: {
        "name": "Yuri Cup Season 6",
        "file": _resolve_file("Yuri Cup Season 6 Probably.xlsx"),
        "sheet": "Coach Stats",
        "fmt": "coach_stats",
    },
}

# Coach name aliases: alias.lower() → canonical display name
# All aliases for the same person must resolve to the same canonical name.
COACH_ALIASES = {
    "kakob": "Jacob",   # S6 match-log spelling
    "beckham": "Jacob", # S8 name for the same person
}

# Tier labels used in the Draft sheet
TIER_LABELS = {"Tier 1", "Tier 2", "Tier 3", "Tier 4", "Tier 5", "Uber 1", "Uber 2", "Free Pick"}


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _canonical(name: str) -> str:
    """Return the canonical coach name (applying aliases if known)."""
    return COACH_ALIASES.get(name.lower().strip(), name.strip())


def _norm_poke(name: str) -> str:
    """Normalize a pokemon name for matching: strip parenthetical suffixes."""
    return re.sub(r'\s*\(.*?\)', '', name).strip()


def _count_faints(faint_val) -> int:
    """Convert a faint annotation cell to a death count integer.

    'Direct' or 'Passive'        → 1
    'Direct 2' / 'Passive 2'     → 2
    'Direct 3'                   → 3
    'Passive 1 Direct 1'         → 2  (sum of all numeric tokens)
    0 / None / ''                → 0
    """
    if not faint_val or faint_val == 0:
        return 0
    s = str(faint_val).strip()
    if not s or s == "0":
        return 0
    # Simple named cases without a number mean 1
    if s in ("Direct", "Passive"):
        return 1
    # Extract all digit tokens and sum them
    nums = re.findall(r'\d+', s)
    if nums:
        return sum(int(n) for n in nums)
    # Fallback: count occurrences of "direct"/"passive" (each = 1)
    low = s.lower()
    return low.count("direct") + low.count("passive")


# ─── S8 / S7: Draft sheet parser ─────────────────────────────────────────────

def parse_draft_sheet(wb):
    """Parse the 'Draft' sheet for per-coach tier and point assignments.

    Returns dict: (coach_id: int, pokemon_name_lower: str) →
                  {"tier": str, "points": int, "is_free_pick": int}
    """
    if "Draft" not in wb.sheetnames:
        return {}

    ws = wb["Draft"]
    rows = list(ws.iter_rows(values_only=True))

    result = {}
    # col_map: pts_col → (coach_id, pokemon_col)
    # Built fresh each time a "Coach:" header row is encountered.
    col_map = {}

    for row in rows:
        if not row or len(row) < 3:
            continue

        cell2 = str(row[2]).strip() if row[2] is not None else ""

        if cell2 == "Coach:":
            # Rebuild column map from this header row.
            # Pattern: col[j] = int coach_id, col[j+2] = str coach_name
            col_map = {}
            for j in range(3, len(row) - 2):
                v = row[j]
                if not isinstance(v, (int, float)) or v <= 0:
                    continue
                if v != int(v):  # must be a whole number
                    continue
                name_col = j + 2
                if name_col >= len(row):
                    continue
                name_val = row[name_col]
                if isinstance(name_val, str) and name_val.strip():
                    col_map[j] = (int(v), name_col)  # pts_col → (coach_id, pokemon_col)

        elif cell2 in TIER_LABELS:
            is_fp = 1 if cell2 == "Free Pick" else 0
            for pts_col, (coach_id, poke_col) in col_map.items():
                if pts_col >= len(row) or poke_col >= len(row):
                    continue
                pts_val = row[pts_col]
                poke_val = row[poke_col]
                if not isinstance(poke_val, str) or not poke_val.strip():
                    continue
                pname = poke_val.strip()
                pts = int(pts_val) if isinstance(pts_val, (int, float)) else 0
                result[(coach_id, pname.lower())] = {
                    "tier": cell2,
                    "points": pts,
                    "is_free_pick": is_fp,
                }

    return result


# ─── S6: Tier List + Rosters parsers ─────────────────────────────────────────

def _parse_s6_rosters(wb, abbrev_to_cid):
    """Parse S6's 'Rosters' sheet for definitive per-coach pokemon ownership.

    The sheet stacks 4 teams per column group (base cols 1, 5, 9, 13).
    Each section is separated by a 'Forfeit' row; sub-sections have an
    abbreviation in col base+2 and end with a 'Pokemon' header before picks.

    Returns dict: (coach_id: int, pokemon_name_lower: str) → True
    """
    if "Rosters" not in wb.sheetnames:
        return {}

    ws = wb["Rosters"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return {}

    header_row = rows[2]  # row 2 has team names and abbreviations
    # Find base columns: pattern repeats every 4 cols starting at 1
    base_cols = []
    j = 1
    while j + 2 < len(header_row):
        abbrev = header_row[j + 2]
        if isinstance(abbrev, str) and abbrev.strip() and len(abbrev.strip()) <= 6 and abbrev.strip() in abbrev_to_cid:
            base_cols.append(j)
        j += 4

    ownership = {}
    for base in base_cols:
        # First section team from header row 2
        current_cid = abbrev_to_cid.get(str(header_row[base + 2]).strip())
        in_pokemon = False

        for row_i, row in enumerate(rows):
            if row_i < 7:  # skip meta rows; row 7 has the 'Pokemon' header for first section
                continue
            if not row or base >= len(row):
                continue
            val = row[base]
            abbrev_here = row[base + 2] if base + 2 < len(row) else None

            if isinstance(val, str) and val.strip() == "Forfeit":
                in_pokemon = False
                continue

            # New sub-section detected by abbreviation in col base+2
            if isinstance(abbrev_here, str) and abbrev_here.strip() and len(abbrev_here.strip()) <= 6:
                cid_new = abbrev_to_cid.get(abbrev_here.strip())
                if cid_new is not None:
                    current_cid = cid_new
                    in_pokemon = False
                    continue

            if isinstance(val, str) and val.strip() == "Pokemon":
                in_pokemon = True
                continue

            if in_pokemon and isinstance(val, str) and val.strip():
                if current_cid is not None:
                    ownership[(current_cid, _norm_poke(val.strip()).lower())] = True

    return ownership


def _parse_s6_points(wb, abbrev_to_cid):
    """Return (coach_id, pokemon_lower) → point_value using Rosters for ownership
    and Tier List for draft costs.

    Tier List column-header values (22, 21, …, 1) are the point costs.
    Rosters tab gives definitive ownership (includes post-draft trades).
    """
    # 1. Build pokemon_lower → pts from Tier List (ignore team column here)
    tier_pts = {}
    if "Tier List" in wb.sheetnames:
        ws_tl = wb["Tier List"]
        tl_rows = list(ws_tl.iter_rows(values_only=True))
        if len(tl_rows) >= 2:
            header = tl_rows[1]
            col_to_pts = {j: int(v) for j, v in enumerate(header)
                          if isinstance(v, (int, float)) and 1 <= int(v) <= 30}
            for row in tl_rows[2:]:
                if not row:
                    continue
                for h, pts in col_to_pts.items():
                    if h >= len(row):
                        continue
                    poke = row[h]
                    team = row[h + 1] if h + 1 < len(row) else None
                    if not isinstance(poke, str) or not poke.strip():
                        continue
                    if not isinstance(team, str) or not team.strip() or team.strip() == "FREE":
                        continue
                    pkey = _norm_poke(poke.strip()).lower()
                    if pkey not in tier_pts:
                        tier_pts[pkey] = pts  # first occurrence wins

    # 2. Get definitive ownership from Rosters tab
    ownership = _parse_s6_rosters(wb, abbrev_to_cid)

    # 3. Combine: assign tier cost to each owned pokemon (0 if not in Tier List = free pick)
    result = {}
    for (cid, pname_lower) in ownership:
        result[(cid, pname_lower)] = tier_pts.get(pname_lower, 0)
    return result


# ─── S8 / S7  ────────────────────────────────────────────────────────────────

def parse_data_sheet(ws, wb=None):
    rows = list(ws.iter_rows(values_only=True))

    # ── 1. Coaches ────────────────────────────────────────────────────────────
    coaches = []
    id_to_coach = {}  # id → {id, coach_name, team_name, ...}

    for row in rows[1:]:
        if not row or len(row) < 8:
            continue
        # Stop when col[1] is no longer a team-ID number
        if not isinstance(row[1], (int, float)) or row[1] is None:
            break
        # Pool A
        cid_a = int(row[1])
        cname_a = _canonical(str(row[2]).strip()) if row[2] else ""
        tname_a = str(row[3]).strip() if row[3] else cname_a
        if cname_a:
            entry = {"id": cid_a, "coach_name": cname_a, "team_name": tname_a,
                     "color": "#888", "logo_url": "", "pool": "A"}
            coaches.append(entry)
            id_to_coach[cid_a] = entry
        # Pool B
        if isinstance(row[5], (int, float)) and row[5] is not None and row[6]:
            cid_b = int(row[5])
            cname_b = _canonical(str(row[6]).strip()) if row[6] else ""
            tname_b = str(row[7]).strip() if row[7] else cname_b
            if cname_b:
                entry_b = {"id": cid_b, "coach_name": cname_b, "team_name": tname_b,
                           "color": "#888", "logo_url": "", "pool": "B"}
                coaches.append(entry_b)
                id_to_coach[cid_b] = entry_b

    # Build lookup including aliases (e.g. "beckham" → ID of "Jacob")
    name_to_id = {}
    for c in coaches:
        name_to_id[c["coach_name"].lower()] = c["id"]
        for alias, canon in COACH_ALIASES.items():
            if canon.lower() == c["coach_name"].lower():
                name_to_id[alias] = c["id"]

    # ── 2. Schedule ───────────────────────────────────────────────────────────
    # Each match appears twice (once from each coach's POV). Deduplicate by
    # (week, frozenset of coach IDs) — keep the first occurrence.
    schedule = []
    seen_matches = set()
    for row in rows[1:]:
        if not row or len(row) < 24:
            continue
        if row[17] != "vs.":
            continue
        coach1_name = str(row[14]).strip() if row[14] else ""
        if not coach1_name or coach1_name == "Coach Name":
            continue
        c1_id = row[13]
        c2_id = row[21]
        if c1_id is None or c2_id is None:
            continue
        c1_id = int(c1_id)
        c2_id = int(c2_id)
        score1 = row[16]
        score2 = row[18]
        if score1 is None or score2 is None:
            continue
        week = int(row[11]) if isinstance(row[11], (int, float)) else row[11]
        match_key = (week, frozenset((c1_id, c2_id)))
        if match_key in seen_matches:
            continue
        seen_matches.add(match_key)
        schedule.append({
            "week": week,
            "coach1_id": c1_id,
            "coach2_id": c2_id,
            "score1": float(score1),
            "score2": float(score2),
            "diff": row[23],
        })

    # ── 3. Per-pokemon stats & roster ─────────────────────────────────────────
    match_stats = []
    roster_set = {}  # (coach_id, pokemon_name) → True
    draft_tiers = {}  # pokemon_name.lower() → {name, type1, type2, spe, points}

    for row in rows[1:]:
        if not row or len(row) < 101:
            continue
        if not isinstance(row[90], (int, float)):
            continue
        coach_name = str(row[91]).strip() if row[91] else ""
        pokemon_name = str(row[93]).strip() if row[93] else ""
        if not coach_name or not pokemon_name:
            continue
        cid = name_to_id.get(coach_name.lower())
        if cid is None:
            print(f"  [WARN] Unknown coach '{coach_name}' in pokemon stats — skipping")
            continue
        kills = float(row[98] or 0)
        deaths = float(row[99] or 0)
        match_stats.append({
            "coach_id": cid,
            "pokemon_name": pokemon_name,
            "kills": kills,
            "deaths": deaths,
        })
        roster_set[(cid, pokemon_name)] = True
        # Collect type/speed info for draft_tiers lookup
        t1 = str(row[94]).strip() if row[94] and str(row[94]).strip() not in ("-", "") else ""
        t2 = str(row[95]).strip() if row[95] and str(row[95]).strip() not in ("-", "") else ""
        spe = int(row[96]) if isinstance(row[96], (int, float)) and row[96] else 0
        pkey = pokemon_name.lower()
        if pkey not in draft_tiers:
            draft_tiers[pkey] = {"name": pokemon_name, "type1": t1, "type2": t2, "spe": spe, "points": 0}

    # ── 4. Merge Draft sheet point/tier data ──────────────────────────────────
    draft_map = parse_draft_sheet(wb) if wb is not None else {}
    if draft_map:
        # Update pokemon_roster with tier + point values
        pokemon_roster = []
        for (cid, pname) in roster_set:
            info = draft_map.get((cid, pname.lower()), {})
            pokemon_roster.append({
                "coach_id": cid,
                "pokemon_name": pname,
                "tier": info.get("tier"),
                "points": info.get("points", 0),
                "is_free_pick": info.get("is_free_pick", 0),
            })
        # Update draft_tiers points from first matching Draft entry
        for pkey, d in draft_tiers.items():
            for (cid2, pname_l), info in draft_map.items():
                if pname_l == pkey:
                    d["points"] = info["points"]
                    break
        print(f"  Draft map   : {len(draft_map)} entries merged")
    else:
        pokemon_roster = [
            {"coach_id": cid, "pokemon_name": pname, "tier": None, "points": 0, "is_free_pick": 0}
            for (cid, pname) in roster_set
        ]

    return {
        "coaches": coaches,
        "schedule": schedule,
        "match_stats": match_stats,
        "pokemon_roster": pokemon_roster,
        "draft_tiers": list(draft_tiers.values()),
        "match_games": [],
        "rules": {},
        "settings": {},
        "transactions": [],
    }


# ─── S6  ──────────────────────────────────────────────────────────────────────

def parse_coach_stats_sheet(ws, wb=None):
    rows = list(ws.iter_rows(values_only=True))

    # ── 1. Coaches from team rows (rows 1-16) ─────────────────────────────────
    coaches = []
    team_to_coach_id = {}   # team_name.lower() → coach_id
    abbrev_to_coach_id = {} # team_abbrev → coach_id  (for Tier List lookup)

    for i, row in enumerate(rows[1:], 1):
        if not row or len(row) < 5:
            break
        abbrev = row[0]
        team_name = row[1]
        coach_name = row[2]
        wins = row[3]
        losses = row[4]
        if not abbrev or not isinstance(abbrev, str) or not team_name or not coach_name:
            break
        if abbrev == "-":
            continue
        cid = i  # use row index as ID (1-16)
        entry = {
            "id": cid,
            "coach_name": _canonical(str(coach_name).strip()),
            "team_name": str(team_name).strip(),
            "color": "#888",
            "logo_url": "",
            "pool": "A",
            "wins": int(wins) if isinstance(wins, (int, float)) else 0,
            "losses": int(losses) if isinstance(losses, (int, float)) else 0,
        }
        coaches.append(entry)
        team_to_coach_id[str(team_name).strip().lower()] = cid
        abbrev_to_coach_id[str(abbrev).strip()] = cid

    # Build lookup including all aliases → coach ID
    coach_name_to_id = {}
    for c in coaches:
        coach_name_to_id[c["coach_name"].lower()] = c["id"]
        for alias, canon in COACH_ALIASES.items():
            if canon.lower() == c["coach_name"].lower():
                coach_name_to_id[alias] = c["id"]

    # ── 2. Per-pokemon summary (cols 1-8, rows 21+) ───────────────────────────
    match_stats = []
    roster_set = {}

    for row in rows[21:]:
        if not row or len(row) < 8:
            continue
        pokemon_name = row[1]
        if not pokemon_name or not isinstance(pokemon_name, str) or pokemon_name == "Pokemon":
            continue
        team_name = row[7]
        if not team_name or not isinstance(team_name, str):
            continue
        cid = team_to_coach_id.get(team_name.strip().lower())
        if cid is None:
            continue
        direct_kos = float(row[2] or 0)
        passive_kos = float(row[3] or 0)
        kills = direct_kos + passive_kos
        pname_clean = _norm_poke(pokemon_name.strip())
        match_stats.append({
            "coach_id": cid,
            "pokemon_name": pname_clean,
            "kills": kills,
            "deaths": 0.0,
        })
        roster_set[(cid, pname_clean)] = True

    # ── 3. Faint tracking (cols 12, 14, 17) ──────────────────────────────────
    # col[12] = owner coach,  col[14] = pokemon that fainted,  col[17] = faint type
    death_stats = {}  # (coach_id, pokemon_name) → total_deaths

    for row in rows[1:]:
        if not row or len(row) < 18:
            continue
        faint_val = row[17]
        if not faint_val or faint_val == 0:
            continue
        n_deaths = _count_faints(faint_val)
        if n_deaths == 0:
            continue
        owner_raw = str(row[12]).strip() if row[12] else ""
        victim_raw = _norm_poke(str(row[14]).strip()) if row[14] else ""
        if not owner_raw or not victim_raw or victim_raw in ("#N/A", "0"):
            continue
        cid = coach_name_to_id.get(owner_raw.lower())
        if cid is None:
            continue
        key = (cid, victim_raw)
        death_stats[key] = death_stats.get(key, 0) + n_deaths

    # Merge death counts into match_stats; add missing entries (0 kills, N deaths)
    ms_lookup = {(m["coach_id"], m["pokemon_name"]): m for m in match_stats}
    for (cid, pname), deaths in death_stats.items():
        if (cid, pname) in ms_lookup:
            ms_lookup[(cid, pname)]["deaths"] = float(deaths)
        else:
            # Pokemon fainted but had 0 kills — add it
            entry = {"coach_id": cid, "pokemon_name": pname, "kills": 0.0, "deaths": float(deaths)}
            match_stats.append(entry)
            ms_lookup[(cid, pname)] = entry
            roster_set[(cid, pname)] = True

    # ── 4. Schedule from match log (col11=week, col20=Win/Loss, col21=diff) ───
    # Each row in the match log gives: coach (col12) vs opponent (col18), result, diff
    # Collect all "Win" sides to reconstruct matches without duplicates.
    seen_pairs = set()
    schedule = []

    for row in rows[1:]:
        if not row or len(row) < 22:
            continue
        week = row[11]
        if not isinstance(week, (int, float)):
            continue
        week = int(week)
        coach_name = str(row[12]).strip() if row[12] else ""
        opponent_name = str(row[18]).strip() if row[18] else ""
        result = str(row[20]).strip() if row[20] else ""
        diff = row[21]
        if not coach_name or not opponent_name or result not in ("Win", "Loss"):
            continue
        cid = coach_name_to_id.get(coach_name.lower())
        oid = coach_name_to_id.get(opponent_name.lower())
        if cid is None or oid is None:
            continue
        pair_key = (week, min(cid, oid), max(cid, oid))
        if pair_key in seen_pairs:
            continue
        seen_pairs.add(pair_key)
        # Winner gets score1 > 0, loser gets score2 = 0.
        # Use 1 as minimum to ensure 0-differential wins still count.
        diff_val = float(diff) if isinstance(diff, (int, float)) else 0.0
        score1 = float(abs(diff_val)) or 1.0
        if result == "Win":
            c1, c2 = cid, oid
        else:
            c1, c2 = oid, cid   # oid = winner
        schedule.append({
            "week": week,
            "coach1_id": c1,
            "coach2_id": c2,
            "score1": score1,
            "score2": 0.0,
            "diff": abs(diff_val),
        })

    # ── 5. Pokemon roster + optional S6 point values ──────────────────────────
    if wb is not None:
        pts_map = _parse_s6_points(wb, abbrev_to_coach_id)
    else:
        pts_map = {}

    if pts_map:
        print(f"  S6 pts map  : {len(pts_map)} entries merged")

    pokemon_roster = [
        {
            "coach_id": cid,
            "pokemon_name": pname,
            "tier": None,
            "points": pts_map.get((cid, pname.lower()), 0),
            "is_free_pick": 0,
        }
        for (cid, pname) in roster_set
    ]

    return {
        "coaches": coaches,
        "schedule": schedule,
        "match_stats": match_stats,
        "pokemon_roster": pokemon_roster,
        "draft_tiers": [],
        "match_games": [],
        "rules": {},
        "settings": {},
        "transactions": [],
    }


# ─── DB helpers ───────────────────────────────────────────────────────────────

def upsert_season(season_num, name, data):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        # Ensure season_num column exists (mirrors Flask _migrate_db)
        try:
            conn.execute("ALTER TABLE seasons ADD COLUMN season_num INTEGER DEFAULT 0")
            conn.commit()
        except sqlite3.OperationalError:
            pass  # column already exists

        existing = conn.execute(
            "SELECT id FROM seasons WHERE season_num=?", (season_num,)
        ).fetchone()
        data_json = json.dumps(data)
        now = datetime.utcnow().isoformat()
        if existing:
            conn.execute(
                "UPDATE seasons SET name=?, data_json=?, archived_at=? WHERE id=?",
                (name, data_json, now, existing["id"]),
            )
            action = f"updated (id={existing['id']})"
        else:
            conn.execute(
                "INSERT INTO seasons (name, season_num, data_json, archived_at) VALUES (?,?,?,?)",
                (name, season_num, data_json, now),
            )
            action = "inserted"
        conn.commit()
        return action
    finally:
        conn.close()


# ─── Main ─────────────────────────────────────────────────────────────────────

def import_season(season_num):
    cfg = SEASONS[season_num]
    print(f"\n=== Season {season_num}: {cfg['name']} ===")
    print(f"  File : {cfg['file']}")
    print(f"  Sheet: {cfg['sheet']}")

    if not os.path.exists(cfg["file"]):
        print(f"  [ERROR] File not found: {cfg['file']}")
        print(f"  Upload the Excel file to {os.path.join(PROJECT_ROOT, 'data', os.path.basename(cfg['file']))}")
        return
    wb = openpyxl.load_workbook(cfg["file"], data_only=True)
    if cfg["sheet"] not in wb.sheetnames:
        print(f"  [ERROR] Sheet '{cfg['sheet']}' not found. Available: {wb.sheetnames}")
        return
    ws = wb[cfg["sheet"]]

    if cfg["fmt"] == "data_sheet":
        data = parse_data_sheet(ws, wb)
    else:
        data = parse_coach_stats_sheet(ws, wb)

    print(f"  Coaches     : {len(data['coaches'])}")
    print(f"  Schedule    : {len(data['schedule'])} matches")
    print(f"  Match stats : {len(data['match_stats'])} entries")
    print(f"  Roster      : {len(data['pokemon_roster'])} entries")
    print(f"  Draft tiers : {len(data['draft_tiers'])} entries")

    action = upsert_season(season_num, cfg["name"], data)
    print(f"  DB          : {action}")


if __name__ == "__main__":
    targets = [int(x) for x in sys.argv[1:] if x.isdigit()] or [8, 7, 6]
    for sn in targets:
        if sn not in SEASONS:
            print(f"[ERROR] Unknown season {sn}. Valid: {list(SEASONS.keys())}")
            continue
        import_season(sn)
    print("\nDone.")
